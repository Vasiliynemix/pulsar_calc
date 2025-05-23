import os
import shutil

import pandas as pd
from loguru import logger
from openpyxl.reader.excel import load_workbook

from src.config import cfg
from src.storage.db.models import Product


class ExcelProvider:
    PRODUCT_COLUMNS = [
        "",
        "Категория",
        "Модель",
        "Хешрейт",
        "Потребление в Вт/ч",
        "Цена в usdt",
        "Алгоритм",
    ]

    def __init__(self):
        pass

    async def create(self, file_path: str, products: list[Product] | None = None):
        # df = pd.DataFrame(columns=self.PRODUCT_COLUMNS)
        #
        # if products:
        #     products_list = []
        #     for i, product in enumerate(products):
        #         products_list.append(
        #             {
        #                 self.PRODUCT_COLUMNS[0]: i + 1,
        #                 self.PRODUCT_COLUMNS[1]: product.category,
        #                 self.PRODUCT_COLUMNS[2]: product.name,
        #                 self.PRODUCT_COLUMNS[3]: product.terahesh,
        #                 self.PRODUCT_COLUMNS[4]: product.consumption,
        #                 self.PRODUCT_COLUMNS[5]: product.price,
        #                 self.PRODUCT_COLUMNS[6]: product.algorithm,
        #             },
        #         )
        #     df = pd.DataFrame.from_records(products_list)
        #
        # writer = pd.ExcelWriter(file_path, engine="xlsxwriter")
        # df.to_excel(writer, sheet_name="Продукты", index=False)
        # writer.close()

        table_path = shutil.copy(
            src=cfg.paths.products_template_path,
            dst=file_path,
        )

        book = load_workbook(filename=table_path)
        sheet = book["products"]

        if products is not None:
            for row, product in enumerate(products, start=2):
                sheet.cell(row=row, column=1, value=f"{row - 1}")
                sheet.cell(row=row, column=2, value=f"{product.category}")
                sheet.cell(row=row, column=3, value=f"{product.name}")
                sheet.cell(row=row, column=4, value=f"{self.__to_int_if_possible(product.terahesh)}")
                sheet.cell(row=row, column=5, value=f"{self.__to_int_if_possible(product.consumption)}")
                sheet.cell(row=row, column=6, value=f"{self.__to_int_if_possible(product.price)}")
                sheet.cell(row=row, column=7, value=f"{product.algorithm}")

        book.save(table_path)

    async def parse(self, file_path: str) -> list[Product] | str | None:
        try:
            df = pd.read_excel(file_path)  # noqa
            products = []
            for index, row in df.iterrows():
                try:
                    if (
                            pd.isnull(row[self.PRODUCT_COLUMNS[1]])
                            and pd.isnull(row[self.PRODUCT_COLUMNS[2]])
                            and pd.isnull(row[self.PRODUCT_COLUMNS[3]])
                            and pd.isnull(row[self.PRODUCT_COLUMNS[4]])
                            and pd.isnull(row[self.PRODUCT_COLUMNS[6]])
                            and pd.isnull(row[self.PRODUCT_COLUMNS[6]])
                    ):
                        continue

                    if pd.isnull(row[self.PRODUCT_COLUMNS[1]]):
                        raise ValueError
                    if pd.isnull(row[self.PRODUCT_COLUMNS[2]]):
                        raise ValueError
                    row[self.PRODUCT_COLUMNS[3]] = float(row[self.PRODUCT_COLUMNS[3]])
                    row[self.PRODUCT_COLUMNS[4]] = float(row[self.PRODUCT_COLUMNS[4]])
                    row[self.PRODUCT_COLUMNS[5]] = float(row[self.PRODUCT_COLUMNS[5]])
                    if pd.isnull(row[self.PRODUCT_COLUMNS[6]]):
                        raise ValueError
                except ValueError:
                    logger.warning(f"ValueError while parsing excel file")
                    error_msg = ""
                    for i, col in enumerate(self.PRODUCT_COLUMNS[1:], start=1):
                        error_msg += f"{col}: {row.to_list()[i]}\n"
                    return (
                        f"В строке {index + 1} ошибка\n\n"  # noqa
                        f"Неверный формат данных.\n\n"
                        f"{error_msg}"
                    )
                product = Product(
                    category=row[self.PRODUCT_COLUMNS[1]],
                    name=row[self.PRODUCT_COLUMNS[2]],
                    terahesh=row[self.PRODUCT_COLUMNS[3]],
                    consumption=row[self.PRODUCT_COLUMNS[4]],
                    price=row[self.PRODUCT_COLUMNS[5]],
                    algorithm=row[self.PRODUCT_COLUMNS[6]],
                )
                products.append(product)
            return products
        except Exception as e:
            logger.error(f"Error while parsing excel file: {e}")
            return None

    @staticmethod
    async def delete(file_path: str):
        if os.path.exists(file_path):
            os.remove(file_path)

    @staticmethod
    def __to_int_if_possible(value: float) -> int | float:
        if value.is_integer():
            return int(value)
        return value
